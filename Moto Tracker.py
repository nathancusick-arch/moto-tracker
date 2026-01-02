import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment, Font

# Rich text support
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
    RICH_TEXT_AVAILABLE = True
except Exception:
    RICH_TEXT_AVAILABLE = False

st.title("Moto Audit Tracker Mapper")

st.write(
        """
        1. Export full month(s) of ALL Moto data (EVERYTHING, INCLUDING incomplete audits)
        2. Drop the file in the below box, it should then give you the output file in your downloads
        3. Paste into the Moto Tracker
        4. Done.
        """
)

# ---------------------------------------------------------
# EXACT SITE ORDER FROM TRACKER
# ---------------------------------------------------------
TRACKER_SITE_ORDER = [
    "SITE32718", "SITE32719", "SITE32720", "SITE32721", "SITE32722",
    "SITE32723", "SITE32724", "SITE32725", "SITE32727", "SITE32728",
    "SITE32729", "SITE32730", "SITE32731", "SITE32732", "SITE32733",
    "SITE32734", "SITE32736", "SITE32737", "SITE32738", "SITE32739",
    "SITE32740", "SITE32741", "SITE32742", "SITE32743", "SITE32744",
    "SITE32745", "SITE32746", "SITE32747", "SITE32748", "SITE32749",
    "SITE32750", "SITE32751", "SITE32752", "SITE32753", "SITE32754",
    "SITE32755", "SITE32756", "SITE32757", "SITE32758", "SITE32759",
    "SITE32760", "SITE32761", "SITE32762", "SITE32763", "SITE32764",
    "SITE32765", "SITE32767", "SITE32768", "SITE32769", "SITE32771",
    "SITE32772", "SITE32773", "SITE48318", "SITE306813",
]

# ---------------------------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------------------------
def day_ordinal(ts: pd.Timestamp) -> str:
    d = ts.day
    if 11 <= d <= 13:
        suffix = "TH"
    else:
        suffix = {1: "ST", 2: "ND", 3: "RD"}.get(d % 10, "TH")
    return f"{d}{suffix}"

def _month_base_label(ts: pd.Timestamp) -> str:
    month_name = ts.strftime("%B")
    year_short = ts.strftime("%y")
    return f"{month_name} '{year_short}"

def _token_group(tok: str) -> str:
    # tok is normalized lower-case
    if tok in {"monthly", "weekly"}:
        return "mw"
    if tok == "random":
        return "random"
    return "other"

def _argb(rgb6: str) -> str:
    rgb6 = (rgb6 or "000000").replace("#", "").upper()
    if len(rgb6) != 6:
        rgb6 = "000000"
    return "FF" + rgb6

def segment_color(row) -> str:
    """Return hex RGB (no #) for the audit segment, per rules."""
    if str(row.get("order_schedule_type_norm", "")).strip().lower() == "emergency":
        return "00B050"  # green (overrides)
    tok = str(row.get("token_class", "")).strip().lower()
    if tok == "weekly":
        return "FF0000"  # red
    if tok == "random":
        return "0070C0"  # blue
    return "000000"      # black

def _idxmin_dt(series: pd.Series) -> int:
    """Pick idx of earliest datetime, handling all-NaT."""
    s = pd.to_datetime(series, errors="coerce")
    if s.notna().any():
        return s.idxmin()
    return s.index[0]

def _is_blank_rich(val) -> bool:
    """Return True if a rich-cell placeholder is blank/NA (safe for lists)."""
    if val is None:
        return True
    # Empty string or empty list
    if val == "" or val == []:
        return True
    # Pandas may store NaN floats in DataFrames
    if isinstance(val, float):
        try:
            return pd.isna(val)
        except Exception:
            return False
    # pd.NA / NaT etc (but avoid applying to lists/dicts)
    try:
        return pd.isna(val) if not isinstance(val, (list, dict)) else (len(val) == 0)
    except Exception:
        return False

# ---------------------------------------------------------
# STREAMLIT FILE UPLOADER
# ---------------------------------------------------------
uploaded_file = st.file_uploader("Upload audits_basic_data_export.csv", type=["csv"])

if uploaded_file is not None:
    # Load CSV
    df = pd.read_csv(uploaded_file)

    # Normalize status; remove only deleted
    df["status_norm"] = df.get("status").astype(str).str.strip().str.lower()
    df = df[df["status_norm"] != "deleted"].copy()

    # Approved vs non-approved
    df["is_approved"] = df["status_norm"] == "approved"

    # Parse dates (UK dd/mm/yyyy). Month assignment:
    # - Approved: month from date_of_visit
    # - Non-approved: month from start_date
    df["date_of_visit_dt"] = pd.to_datetime(df.get("date_of_visit"), dayfirst=True, errors="coerce")
    df["start_date_dt"] = pd.to_datetime(df.get("start_date"), dayfirst=True, errors="coerce")
    df["month_dt"] = df["date_of_visit_dt"].where(df["is_approved"], df["start_date_dt"])
    df = df[df["month_dt"].notna()].copy()

    # Parse time_of_visit for approved tie-breaks
    df["time_of_visit_td"] = pd.to_timedelta(df.get("time_of_visit"), errors="coerce").fillna(pd.Timedelta(0))
    df["visit_dt"] = df["date_of_visit_dt"] + df["time_of_visit_td"]

    # Normalized fields
    df["PRIMARY_RESULT"] = df.get("primary_result").astype(str).str.upper()
    df["token_class"] = df.get("tokens").astype(str).str.strip().str.lower()
    df["token_group"] = df["token_class"].apply(_token_group)

    # Ensure order_schedule_type exists
    if "order_schedule_type" not in df.columns:
        df["order_schedule_type"] = ""
    df["order_schedule_type_norm"] = df["order_schedule_type"].fillna("").astype(str).str.strip().str.lower()

    # Grouping keys by month
    df["col_year"] = df["month_dt"].dt.year
    df["col_month"] = df["month_dt"].dt.month
    df["month_base"] = df["month_dt"].apply(_month_base_label)

    # Sort to make selection deterministic; approved priority is enforced below
    df = df.sort_values(
        ["site_internal_id", "col_year", "col_month", "is_approved", "visit_dt", "month_dt"],
        ascending=[True, True, True, False, True, True],
    )

    # ---------------------------------------------------------
    # Assign base vs extra per site+month
    # ---------------------------------------------------------
    df["is_base"] = False

    for (site, y, mth), g in df.groupby(["site_internal_id", "col_year", "col_month"], sort=False):
        base_idx = None

        g_mw = g[g["token_group"] == "mw"]
        g_mw_approved = g_mw[g_mw["is_approved"]]
        g_mw_nonapproved = g_mw[~g_mw["is_approved"]]

        # 1) Monthly/Weekly takes priority (approved preferred)
        if len(g_mw_approved) > 0:
            if len(g_mw_approved) == 1:
                # Only one approved MW: becomes base even if emergency
                base_idx = _idxmin_dt(g_mw_approved["visit_dt"])
            else:
                # Multiple approved MW: prefer earliest non-emergency
                g_non_em = g_mw_approved[g_mw_approved["order_schedule_type_norm"] != "emergency"]
                if len(g_non_em) > 0:
                    base_idx = _idxmin_dt(g_non_em["visit_dt"])
                else:
                    # All approved MW are emergency: earliest emergency becomes base
                    base_idx = _idxmin_dt(g_mw_approved["visit_dt"])

        elif len(g_mw_nonapproved) > 0:
            # No approved MW: choose among non-approved MW (start_date only)
            if len(g_mw_nonapproved) == 1:
                base_idx = _idxmin_dt(g_mw_nonapproved["month_dt"])
            else:
                g_non_em = g_mw_nonapproved[g_mw_nonapproved["order_schedule_type_norm"] != "emergency"]
                if len(g_non_em) > 0:
                    base_idx = _idxmin_dt(g_non_em["month_dt"])
                else:
                    base_idx = _idxmin_dt(g_mw_nonapproved["month_dt"])

        else:
            # 2) No Monthly/Weekly: earliest Random becomes base (approved preferred)
            g_rand = g[g["token_group"] == "random"]
            g_rand_approved = g_rand[g_rand["is_approved"]]
            g_rand_nonapproved = g_rand[~g_rand["is_approved"]]

            if len(g_rand_approved) > 0:
                base_idx = _idxmin_dt(g_rand_approved["visit_dt"])
            elif len(g_rand_nonapproved) > 0:
                base_idx = _idxmin_dt(g_rand_nonapproved["month_dt"])
            else:
                base_idx = None

        if base_idx is not None:
            df.loc[base_idx, "is_base"] = True

    df["tracker_column"] = df["month_base"] + df["is_base"].map(lambda b: "" if b else " Extra")

    # Chronological order for writing output strings
    df["sort_dt"] = df["visit_dt"].where(df["is_approved"], df["month_dt"])
    df = df.sort_values(["site_internal_id", "col_year", "col_month", "tracker_column", "sort_dt"])

    # Build dynamic column list based on months present
    date_groups = (
        df[["col_year", "col_month"]]
        .drop_duplicates()
        .sort_values(["col_year", "col_month"])
        .to_records(index=False)
    )

    final_columns = []
    for y, m in date_groups:
        month_name = pd.to_datetime(f"{y}-{m}-01").strftime("%B")
        year_short = str(y)[-2:]
        base = f"{month_name} '{year_short}"
        final_columns.append(base)
        final_columns.append(base + " Extra")

    # ---------------------------------------------------------
    # Build output tables:
    #  - out_text: strings for on-screen preview
    #  - out_rich: per-cell list of segments [{'text':..., 'rgb':...}]
    # ---------------------------------------------------------
    out_text = pd.DataFrame(index=TRACKER_SITE_ORDER, columns=final_columns)
    out_text.index.name = "Site Code"
    out_rich = pd.DataFrame(index=TRACKER_SITE_ORDER, columns=final_columns)
    out_rich.index.name = "Site Code"

    for _, row in df.iterrows():
        site = str(row.get("site_internal_id"))
        col = row.get("tracker_column")

        if site not in out_text.index or col not in out_text.columns:
            continue

        rgb = segment_color(row)

        if bool(row.get("is_approved")):
            # Approved: show result + visit day
            if pd.isna(row.get("date_of_visit_dt")):
                val = str(row.get("PRIMARY_RESULT", "")).strip() or "TBC"
            else:
                day_str = day_ordinal(row["date_of_visit_dt"])
                val = f"{row['PRIMARY_RESULT']} - {day_str}"
        else:
            # Non-approved: always TBC
            val = "TBC"

        # Append rich segments (keep duplicates like 'TBC, TBC')
        existing_segments = out_rich.loc[site, col]
        if isinstance(existing_segments, list):
            existing_segments.append({"text": val, "rgb": rgb})
            out_rich.loc[site, col] = existing_segments
        elif pd.isna(existing_segments) or existing_segments in (None, "", []):
            out_rich.loc[site, col] = [{"text": val, "rgb": rgb}]
        else:
            out_rich.loc[site, col] = [{"text": str(existing_segments), "rgb": "000000"}, {"text": val, "rgb": rgb}]

        # Append preview text
        existing = out_text.loc[site, col]
        out_text.loc[site, col] = val if pd.isna(existing) or existing == "" else f"{existing}, {val}"

    # Fill N/A for past months (months before current calendar month)
    today = datetime.today()
    current_y = int(today.strftime("%y"))
    current_m = int(today.strftime("%m"))

    for col in out_text.columns:
        parts = col.replace(" Extra", "").split(" '")
        month_name = parts[0]
        year_short = int(parts[1])
        month_num = pd.to_datetime(month_name, format="%B").month

        if (year_short < current_y) or (year_short == current_y and month_num < current_m):
            out_text[col] = out_text[col].fillna("N/A")
            for site in out_rich.index:
                if _is_blank_rich(out_rich.loc[site, col]):
                    out_rich.loc[site, col] = "N/A"

    # Preview
    st.subheader("Preview of Output")
    st.dataframe(out_text)

    # ---------------------------------------------------------
    # Prepare XLSX download with rich-text colouring
    # ---------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Moto Tracker Results"

    # Header row
    c0 = ws.cell(row=1, column=1, value="Site Code")
    c0.font = Font(name="Arial", size=10)
    c0.alignment = Alignment(vertical="center")
    for j, col in enumerate(out_text.columns, start=2):
        cj = ws.cell(row=1, column=j, value=col)
        cj.font = Font(name="Arial", size=10)
        cj.alignment = Alignment(vertical="center")

    # Data rows
    for i, site in enumerate(out_text.index, start=2):
        c_site = ws.cell(row=i, column=1, value=site)
        c_site.font = Font(name="Arial", size=10)
        c_site.alignment = Alignment(vertical="center")

        for j, col in enumerate(out_text.columns, start=2):
            cell = ws.cell(row=i, column=j)
            rich_val = out_rich.loc[site, col]

            if rich_val == "N/A":
                cell.value = "N/A"
                cell.alignment = Alignment(vertical="center")
                cell.font = Font(name="Arial", size=10)
            elif isinstance(rich_val, list) and len(rich_val) > 0:
                if RICH_TEXT_AVAILABLE:
                    rt = CellRichText()
                    for k, seg in enumerate(rich_val):
                        seg_text = str(seg.get("text", ""))
                        seg_rgb = str(seg.get("rgb", "000000"))
                        rt.append(
                            TextBlock(
                                InlineFont(name="Arial", size=10, color=Color(rgb=_argb(seg_rgb))),
                                seg_text,
                            )
                        )
                        if k < len(rich_val) - 1:
                            rt.append(
                                TextBlock(
                                    InlineFont(name="Arial", size=10, color=Color(rgb=_argb("000000"))),
                                    ", ",
                                )
                            )
                    cell.value = rt
                else:
                    # Fallback: no rich text support, write plain string
                    cell.value = ", ".join(str(seg.get("text", "")) for seg in rich_val)

                # Alignment applies regardless of rich-text support
                cell.alignment = Alignment(vertical="center")
                # Font applies to non-rich-text cells
                cell.font = Font(name="Arial", size=10)

            else:
                # leave blank
                cell.alignment = Alignment(vertical="center")
                cell.font = Font(name="Arial", size=10)
                pass

    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)

    if not RICH_TEXT_AVAILABLE:
        st.warning(
            "Rich text formatting isn't available in this environment (openpyxl rich_text missing). "
            "The download will be plain text."
        )

    st.download_button(
        label="Download Moto Tracker Results",
        data=xlsx_buffer,
        file_name="Moto Tracker Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
